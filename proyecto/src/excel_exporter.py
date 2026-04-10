"""
Modulo excel_exporter.py - Exporta datos de facturas a Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime


class ExcelExporter:
    """Exporta datos de facturas a archivos Excel"""

    def _formatear_valor(self, valor: Any) -> str:
        """Convierte valores complejos a texto legible para Excel."""
        if valor is None:
            return ""
        if isinstance(valor, (int, float, str)):
            return str(valor)
        if isinstance(valor, list):
            return ", ".join(self._formatear_valor(elemento) for elemento in valor)
        if isinstance(valor, dict):
            return ", ".join(f"{clave}: {self._formatear_valor(elemento)}" for clave, elemento in valor.items())
        return str(valor)

    def generar_nombre_archivo(self, numero_factura: str, tipo: str = "detallado") -> str:
        """Genera nombre del archivo Excel"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_factura = numero_factura.replace("/", "_").replace(" ", "_") if numero_factura else "factura"
        return f"{nombre_factura}_{timestamp}.xlsx"

    def crear_hoja_unica(self, datos_factura: Dict[str, Any], ruta_salida: str) -> bool:
        """Alias para crear_reporte_detallado"""
        return self.crear_reporte_detallado(datos_factura, ruta_salida)

    def crear_reporte_detallado(self, datos_factura: Dict[str, Any], ruta_salida: str) -> bool:
        """
        Crea un Excel detallado para UNA factura
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Factura"

            # Estilos
            titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            titulo_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            encabezado_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            encabezado_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            borde = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

            # Titulo
            ws['A1'] = 'DATOS DE FACTURA'
            ws['A1'].font = titulo_font
            ws['A1'].fill = titulo_fill
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells('A1:B1')
            ws.row_dimensions[1].height = 25

            def escribir_seccion(titulo_seccion: str, campos: List[tuple[str, Any]], fila_inicio: int) -> int:
                fila_actual = fila_inicio
                ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=2)
                celda_titulo = ws.cell(row=fila_actual, column=1)
                celda_titulo.value = titulo_seccion
                celda_titulo.font = encabezado_font
                celda_titulo.fill = encabezado_fill
                celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
                celda_titulo.border = borde
                fila_actual += 1

                for etiqueta, valor in campos:
                    ws[f'A{fila_actual}'] = etiqueta
                    ws[f'A{fila_actual}'].font = encabezado_font
                    ws[f'A{fila_actual}'].fill = encabezado_fill
                    ws[f'A{fila_actual}'].border = borde

                    ws[f'B{fila_actual}'] = self._formatear_valor(valor)
                    ws[f'B{fila_actual}'].border = borde
                    ws[f'B{fila_actual}'].alignment = Alignment(horizontal='left')
                    fila_actual += 1

                return fila_actual + 1

            # Datos
            fila = 3
            fila = escribir_seccion('IDENTIFICACIÓN', [
                ('Numero Factura:', datos_factura.get('numero_factura', 'N/A')),
                ('Fecha Emision:', datos_factura.get('fecha_emision', datos_factura.get('fecha', 'N/A'))),
                ('Fecha Vencimiento:', datos_factura.get('fecha_vencimiento', 'N/A')),
                ('Tipo Factura:', datos_factura.get('tipo_factura', 'FACTURA')),
            ], fila)

            fila = escribir_seccion('PROVEEDOR', [
                ('Empresa:', datos_factura.get('empresa_proveedor', datos_factura.get('empresa', 'N/A'))),
                ('RUC:', datos_factura.get('ruc_empresa', datos_factura.get('ruc', 'N/A'))),
                ('Direccion:', datos_factura.get('direccion_empresa', '')),
                ('Telefono:', datos_factura.get('telefono_empresa', '')),
                ('Email:', datos_factura.get('email_empresa', '')),
            ], fila)

            fila = escribir_seccion('CLIENTE', [
                ('Empresa:', datos_factura.get('cliente_nombre', '')),
                ('RUC:', datos_factura.get('cliente_ruc', '')),
                ('Direccion:', datos_factura.get('cliente_direccion', '')),
            ], fila)

            fila = escribir_seccion('DETALLE', [
                ('Concepto:', datos_factura.get('concepto', '')),
                ('Cantidad:', datos_factura.get('cantidad_detalle', datos_factura.get('cantidad_items', ''))),
                ('Precio Unitario:', datos_factura.get('precio_unitario', '')),
            ], fila)

            fila = escribir_seccion('TOTALES', [
                ('Subtotal:', datos_factura.get('subtotal', '0.00')),
                ('Descuento:', datos_factura.get('descuento', '0.00')),
                ('IGV:', datos_factura.get('igv', '0.00')),
                ('TOTAL:', datos_factura.get('monto_total', datos_factura.get('total', '0.00'))),
            ], fila)

            fila = escribir_seccion('ADICIONAL', [
                ('Moneda:', datos_factura.get('moneda', 'PEN')),
                ('Forma de Pago:', datos_factura.get('forma_pago', '')),
                ('Observaciones:', datos_factura.get('observaciones', '')),
            ], fila)

            fila = escribir_seccion('METADATOS', [
                ('Fecha Procesamiento:', datos_factura.get('fecha_procesamiento', '')),
                ('Archivo Origen:', datos_factura.get('archivo_origen', '')),
            ], fila)

            # Hojа adicional con items cuando existan
            items = datos_factura.get('items', [])
            if isinstance(items, list) and items:
                ws_items = wb.create_sheet('Items')
                encabezados_items = ['Codigo', 'Descripcion', 'Cantidad', 'Unidad', 'Precio Unitario', 'Subtotal', 'Descuento', 'Impuesto', 'Total']
                for col, encabezado in enumerate(encabezados_items, 1):
                    celda = ws_items.cell(row=1, column=col)
                    celda.value = encabezado
                    celda.font = titulo_font
                    celda.fill = titulo_fill
                    celda.border = borde
                    celda.alignment = Alignment(horizontal='center', vertical='center')

                for fila_idx, item in enumerate(items, 2):
                    if isinstance(item, dict):
                        datos_item = item
                    else:
                        datos_item = getattr(item, '__dict__', {})

                    valores_item = [
                        datos_item.get('codigo', ''),
                        datos_item.get('descripcion', ''),
                        datos_item.get('cantidad', ''),
                        datos_item.get('unidad', ''),
                        datos_item.get('precio_unitario', ''),
                        datos_item.get('subtotal', ''),
                        datos_item.get('descuento', ''),
                        datos_item.get('impuesto', ''),
                        datos_item.get('total', ''),
                    ]

                    for col, valor in enumerate(valores_item, 1):
                        celda = ws_items.cell(row=fila_idx, column=col)
                        celda.value = self._formatear_valor(valor)
                        celda.border = borde
                        celda.alignment = Alignment(horizontal='left')

                anchos_items = [15, 40, 12, 10, 15, 12, 12, 12, 12]
                for col, ancho in enumerate(anchos_items, 1):
                    ws_items.column_dimensions[chr(64 + col)].width = ancho

            # Ancho de columnas
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 40

            # Guardar
            wb.save(ruta_salida)
            return True

        except Exception as e:
            print(f"Error creando Excel: {e}")
            return False

    def crear_reporte_lote(self, facturas_lista: List[Dict[str, Any]], ruta_salida: str) -> bool:
        """
        Crea un Excel con multiples facturas en una tabla
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Facturas"

            # Estilos
            titulo_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            titulo_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            borde = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

            # Encabezados
            encabezados = ['Num. Factura', 'Fecha', 'Empresa', 'RUC', 'Subtotal', 'IGV', 'Total']
            for col, encabezado in enumerate(encabezados, 1):
                celda = ws.cell(row=1, column=col)
                celda.value = encabezado
                celda.font = titulo_font
                celda.fill = titulo_fill
                celda.border = borde
                celda.alignment = Alignment(horizontal='center', vertical='center')

            # Datos
            for fila_idx, factura in enumerate(facturas_lista, 2):
                datos_fila = [
                    factura.get('numero_factura', 'N/A'),
                    factura.get('fecha', 'N/A'),
                    factura.get('empresa', 'N/A'),
                    factura.get('ruc', 'N/A'),
                    factura.get('subtotal', '0.00'),
                    factura.get('igv', '0.00'),
                    factura.get('total', '0.00'),
                ]

                for col, valor in enumerate(datos_fila, 1):
                    celda = ws.cell(row=fila_idx, column=col)
                    celda.value = valor
                    celda.border = borde
                    celda.alignment = Alignment(horizontal='left')

            # Ancho de columnas
            anchos = [15, 12, 25, 15, 12, 10, 12]
            for col, ancho in enumerate(anchos, 1):
                ws.column_dimensions[chr(64 + col)].width = ancho

            # Guardar
            wb.save(ruta_salida)
            return True

        except Exception as e:
            print(f"Error creando Excel lote: {e}")
            return False
